from flask import Flask, redirect, request, make_response, abort, render_template, Response, send_file
from logging.config import dictConfig
from werkzeug.exceptions import BadRequest
import re
import json
import numpy as np
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from pymongo import MongoClient

from .auth import test_auth, no_auth
from .app import app


def choose(config):
    prob = np.array(config['prob'])
    prob = prob / prob.sum()
    idx  = np.random.choice(np.arange(len(prob)), p = prob)
    return config['item'][idx], config['prce'][idx], config['num'][idx]

def to_str(p):
    name, _, num = p
    if num == 1:
        return name
    elif num < 1:
        return ''
    else:
        return f'{name}(x{num})'


@app.route("/random/bill/list", methods=['GET'])
def get_random_bill_list():
    """
    /random/bill json example

    {
        "number": 20,
        "soup": {
            "item": ["豆浆", "青菜粥", "皮蛋瘦肉粥", "豆腐花", "绿豆粥"],
            "prce": [2.5, 3, 4, 4, 4],
            "prob": [9, 7, 3, 3, 3],
            "num": [1, 1, 1, 1, 1]
        },
        "food": {
            "item": ['包子', '油条', '小笼包', '炒面', '生煎锅贴', '煎饼果子'],
            "prce": [2.5, 2.5, 6, 6, 1.5, 9],
            "prob": [8, 8, 5, 3, 5, 1],
            "num": [2, 2, 1, 1, 6, 1]
        }
        ...
    }

    """

    if request.headers.get("Content-Type", default = "") != "application/json":
        abort(Response(response=json.dumps(dict(code = 2, message = "Not found", data = None)), status = 404))

    params = request.get_json()

    if not "token" in params:
        return no_auth
    client = MongoClient(app.config['MONGO_URI'])
    right = test_auth(client['rndsvr']['token'], params['token'], 'anonymous')
    if right is not None:
        return right

    for name in ["number", "soup", "food", "pick", "drink"]:
        if name not in params:
            abort(Response(response=json.dumps(dict(code = 6, message = f"parameter {name} missing", data = None)), status = 400))

    for item in ["soup", "food", "pick", "drink"]:
        for name in ["item", "prce", "prob", "num"]:
            if name not in params[item]:
                abort(Response(response=json.dumps(dict(code = 6, message = f"parameter {name} missing", data = params[item])), status = 400))
        if not len(params[item]['item']) == len(params[item]['prce']) == len(params[item]['prob']) == len(params[item]['num']):
            abort(Response(response=json.dumps(dict(code = 6, message = f"parameter item length not same", data = params[item])), status = 400))

    items = []
    total = 0

    for i in range(params['number']):
        soup  = choose(params['soup'])
        food  = choose(params['food'])
        pick  = choose(params['pick'])
        drink = choose(params['drink'])
        prce = soup[1] * soup[2] + food[1] * food[2] + pick[1] * pick[2] + drink[1] * drink[2]
        items.append(dict(item = f'{to_str(soup)} {to_str(food)} {to_str(pick)} {to_str(drink)}', prce = prce))
        total += prce

    return Response(json.dumps(dict(code = 0, message = "success", data = dict(items = items, total = total))), mimetype = 'application/json')

@app.route("/random/bill/date", methods=['GET'])
def get_random_bill_date():
    """
    {
        "begin": "yyyy-MM-dd",
        "end":   "yyyy-MM-dd",
        "day": "default",
        "men": [
            {
                "name": "john",
                "begin": "yyyy-MM-dd",
                "end":   "yyyy-MM-dd",
            }
            ...
        ]
    }
    """
    if request.headers.get("Content-Type", default = "") != "application/json":
        abort(Response(response=json.dumps(dict(code = 2, message = "Not found", data = None)), status = 404))

    params = request.get_json()

    if not "token" in params:
        return abort(no_auth)
    client = MongoClient(app.config['MONGO_URI'])
    right = test_auth(client['rndsvr']['token'], params['token'], 'anonymous')
    if right is not None:
        return right

    for name in ["begin", "end", "day", "men"]:
        if name not in params:
            abort(Response(response=json.dumps(dict(code = 6, message = f"parameter {name} missing", data = None)), status = 400))

    for item in params["men"]:
        for name in ["name", "begin", "end"]:
            if name not in item:
                abort(Response(response=json.dumps(dict(code = 6, message = f"parameter {name} missing", data = item)), status = 400))

    begin_date = datetime.strptime(params["begin"], '%Y-%m-%d').date()
    end_date   = datetime.strptime(params["end"  ], '%Y-%m-%d').date()

    items = []

    current = begin_date
    count = 0
    while current <= end_date:
        if current.weekday() >= 5:
            current += timedelta(days = 1)
        today = dict(date = str(current), items = [])
        for m in params['men']:
            if datetime.strptime(m["begin"], '%Y-%m-%d').date() <= current <= datetime.strptime(m["end"], '%Y-%m-%d').date():
                today['items'].append(dict(name = m['name']))
                count += 1
        if len(today['items']) > 0:
            items.append(today)
        current += timedelta(days = 1)
    return Response(json.dumps(dict(code = 0, message = "success", data = dict(items = items, count = count))), mimetype = 'application/json')

@app.route("/random/bill/file", methods=['GET'])
def get_random_bill_file():
    """
    {
        "days": [
            {
                "date": "2019-12-20",
                "items": [
                    {
                        "name": "john",
                        "item": "xxxx",
                        "prce": 10
                    },
                    {
                        "name": "who",
                        "item": "xxxx",
                        "prce": 10
                    }
                ]
            },
            {
                "date": "2019-12-22",
                "items": [
                    {
                        "name": "john",
                        "item": "xxxx",
                        "prce": 10
                    },
                    {
                        "name": "who"
                        "item": "xxxx",
                        "prce": 10
                    }
                ]
            }
            ]
    }
    """

    if request.headers.get("Content-Type", default = "") != "application/json":
        abort(Response(response=json.dumps(dict(code = 2, message = "Not found", data = None)), status = 404))

    params = request.get_json()

    if not "token" in params:
        return no_auth
    client = MongoClient(app.config['MONGO_URI'])
    right = test_auth(client['rndsvr']['token'], params['token'], 'anonymous')
    if right is not None:
        return right

    days = params['days']

    for day in days:
        for name in ["date", "items"]:
            if not name in day:
                abort(Response(response=json.dumps(dict(code = 6, message = f"parameter {name} missing", data = day)), status = 400))
        for item in day['items']:
            for name in ['name', 'item', 'prce']:
                if not name in item:
                    abort(Response(response=json.dumps(dict(code = 6, message = f"parameter {name} missing", data = item)), status = 400))


    wb = Workbook()
    ws = wb.active

    # styles
    hl_style_font = Font(name = '宋体', sz = 10, bold = True)
    dt_style_font = Font(name = '宋体', sz = 10)
    center_algn = Alignment(horizontal="center", vertical="center")
    right_algn = Alignment(horizontal="right", vertical="center")

    def fill_cell(col, row, value, font = None, algn = None, fmt = None):
        cell = ws.cell(column = col, row = row, value = value)
        if fmt:
            cell.number_format = fmt
        if font:
            cell.font = font
        if algn:
            cell.alignment = algn
        return cell


    # head line
    fill_cell(1, 1, '日期', font= hl_style_font, algn=center_algn)
    fill_cell(2, 1, '姓名', font= hl_style_font, algn=center_algn)
    fill_cell(3, 1, '种类', font= hl_style_font, algn=center_algn)
    fill_cell(4, 1, '金额', font= hl_style_font, algn=center_algn)
    fill_cell(5, 1, '总额', font= hl_style_font, algn=center_algn)

    line = 2
    for day in days:
        date = day['date']
        begin_line = line
        total = 0
        for item in day['items']:
            fill_cell(2, line, item['name'], font = dt_style_font, algn = center_algn)
            fill_cell(3, line, item['item'], font = dt_style_font, algn = center_algn)
            fill_cell(4, line, item['prce'], font = dt_style_font, algn = center_algn, fmt = '[$￥-804]#,##0.00;[RED]-[$￥-804]#,##0.00')
            total += item['prce']
            line += 1
        ws.merge_cells(
            start_row = begin_line, end_row = line - 1,
            start_column = 1, end_column = 1,
            )
        fill_cell(1, begin_line, day['date'], font = dt_style_font, algn = center_algn, fmt = 'YYYY-M-D')
        ws.merge_cells(
            start_row = begin_line, end_row = line - 1,
            start_column = 5, end_column = 5,
            )

        fill_cell(5, begin_line, f"=sum(D{begin_line}:D{line - 1})", font = dt_style_font, algn = center_algn, fmt = '[$￥-804]#,##0.00;[RED]-[$￥-804]#,##0.00')

    ws.merge_cells(
        start_row = line, end_row = line,
        start_column = 1, end_column = 4
    )
    fill_cell(1, line, '总计', font = hl_style_font, algn = right_algn)
    fill_cell(5, line, f'=sum(D2:D{line-1})', font = dt_style_font, algn = center_algn, fmt = '[$￥-804]#,##0.00;[RED]-[$￥-804]#,##0.00')


    ws.column_dimensions['A'].auto_size = True
    ws.column_dimensions['B'].auto_size = True
    ws.column_dimensions['C'].width     = 60 if not 'c-width' in params else params['c-width']
    ws.column_dimensions['D'].auto_size = True
    ws.column_dimensions['E'].auto_size = True

    f = BytesIO()
    wb.save(f)
    f.seek(0)
    return send_file(f, mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

